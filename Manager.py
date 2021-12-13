import openpyxl
import hashlib

class Manager ():
    def __init__(self):
        self.wb = openpyxl.load_workbook(r'WorkCycle.xlsx')
        self.wbpass= openpyxl.load_workbook(r'pass.xlsx')
        self.ws_names = self.wb.sheetnames
        self.ws_names_to_show = list()
        for i in range(len(self.ws_names)):
            if (i % 2) == 0:
                self.ws_names_to_show.append(self.ws_names[i])
        print('비밀번호를 입력해주세요 :',end ='')
        password  = input(" ")
        Hashed_password = self.Hash_String(password)
        self.SecurityCheck(Hashed_password)

    def Hash_String(self, password : str):
        password = password.encode("UTF-8")
        MD5=hashlib.md5()
        MD5.update(password)
        Enctext=MD5.hexdigest()
        Enctext=Enctext.upper()
        return Enctext

    def SecurityCheck(self, password : str):
        cnt=1
        self.pass_list=list()
        self.passSheet = self.wbpass['Sheet1']
        for i in self.passSheet.rows:
            self.pass_list.append(i[0].value)

        while self.Find_First_Index(self.pass_list,password)==-1:
            if(cnt<5):
                print('비밀번호를 다시 확인해주세요. 5회 오류시 프로그램이 종료됩니다.\n 현재 오류 횟수 :' + str(cnt),end='회')
                password = self.Hash_String(input('\n'))
                cnt = cnt+1
            else:
                print('5회 오류! 프로그램이 종료됩니다.')
                return 0
        print('접속완료!')
        self.Main_Page()


    def Main_Page(self):
        self.is_sheet_selected = False
        self.is_squad_selected = False
        id_to_task = int(input('하고 싶은 작업을 선택하여 숫자로 입력해 주세요 : 0-종료, 1-예정 사역 확인, 2-면제 요청 확인, 3-관리자 추가\n'))
        while (id_to_task <0 or id_to_task >3):
            id_to_task = int(input('입력 에러! 0에서 3사이의 값을 입력해 주세요 : 0-종료, 1-예정 사역 확인, 2-면제 요청 확인, 3-관리자 추가\n'))
        if id_to_task == 1:
            self.Check_Cycle()
        elif id_to_task == 2:
            self.Check_Request()
        elif id_to_task == 3:
            self.Add_manager()
        elif id_to_task == 0:
            return 0

    def Check_Cycle(self):
        len_wsts = len(self.ws_names_to_show)
        print('확인 하고 싶은 사역의 번호를 적어주세요 : ',end='')
        for i in range(len_wsts):
            print(i + 1, end='-')
            if i == len_wsts:
                print(self.ws_names_to_show[i])
            else:
                print(self.ws_names_to_show[i], end=' ')
        self.id_to_search = int(input('\n'))
        while (self.id_to_search < 1 or self.id_to_search > len_wsts):
            self.id_to_search = int(input('입력 에러! 범위에 맞는 숫자 값을 입력해주세요 \n'))
        self.sheet_work = self.wb[self.ws_names_to_show[self.id_to_search - 1]]
        self.sheet_request = self.wb[self.ws_names_to_show[self.id_to_search - 1] + "(수정요청)"]
        self.is_sheet_selected = True
        showIdxList=list()
        for i in self.sheet_work.columns:
            showIdxList.append(i[17].value)
        self.showIdx = self.Find_First_Index(showIdxList,None)
        print(self.ws_names_to_show[self.id_to_search - 1] + " 사이클 현황입니다. ")
        for i in self.sheet_work.rows:
            if(i[0].value==None):
                print(str(i[0].value)+'      ',end ='')
            else:
                print(i[0].value, end=' ')
            for j in range(3):
                if(i[self.showIdx+j].value=='o'):
                    print(i[self.showIdx+j].value,end='    ')
                else:
                    print(i[self.showIdx+j].value,end=' ')
            print()
        next_work = int(input('다음 작업을 선택해주세요 : 0-종료, 1-메인화면, 2-해당사역 면제요청 확인 \n'))
        while (next_work < 0 or next_work > 2):
            next_work = int(input('입력 에러! 범위에 맞는 숫자 값을 입력해주세요 \n'))
        if (next_work == 1):
            self.Main_Page()
        elif (next_work == 2):
            self.Check_Request()
        elif (next_work == 0):
            return 0

    def Check_Request(self):
        if(self.is_sheet_selected == False):
            len_wsts = len(self.ws_names_to_show)
            print('확인 하고 싶은 사역의 번호를 적어주세요 : ', end='')
            for i in range(len_wsts):
                print(i + 1, end='-')
                if i == len_wsts:
                    print(self.ws_names_to_show[i])
                else:
                    print(self.ws_names_to_show[i], end=' ')
            self.id_to_search = int(input('\n'))
            while (self.id_to_search < 1 or self.id_to_search > len_wsts):
                self.id_to_search = int(input('입력 에러! 범위에 맞는 숫자 값을 입력해주세요 \n'))
            self.is_sheet_selected = True
            self.sheet_work = self.wb[self.ws_names_to_show[self.id_to_search - 1]]
            self.sheet_request = self.wb[self.ws_names_to_show[self.id_to_search - 1]+"(수정요청)"]
            showIdxList = list()
            for i in self.sheet_work.columns:
                showIdxList.append(i[17].value)
            self.showIdx = self.Find_First_Index(showIdxList, None)
        self.Squadlist=list()
        for i in self.sheet_work.rows:
            self.Squadlist.append(i[0].value)

        print(self.ws_names_to_show[self.id_to_search - 1]+"에서 다음번 최대 3회차 까지의 면제 요청 분대 리스트와 요청 내용입니다")
        self.asked_squad_list =list()
        self.asked_squad_idx_list=list()
        lists=[0 for i in range(3)]
        lists2=[0 for i in range(3)]
        temp=list()
        for i in range(3):
            for j in self.sheet_request.rows:
                temp.append(j[self.showIdx+i].value)
            lists[i]=temp[17*i:17*i+17]
            if (self.Find_First_Index(lists[i][1:], None) != -1):
                lists2[i] = self.Find_Inds_list(lists[i][1:],None)
            for j in range (16):
                if(self.is_Val_in_list(lists2[i],j)==False and self.is_Val_in_list(self.asked_squad_idx_list,j+1)==False):
                    self.asked_squad_idx_list.append(j+1)
        self.asked_squad_idx_list.sort()
        for i in self.asked_squad_idx_list:                      #면제 요청한 분대의 이름 저장
            self.asked_squad_list.append(self.Squadlist[i])

        self.Squad_Request_list=[[] for i in range(len(self.asked_squad_idx_list))] #분대와 면제 신청 내용 모두 저장하는 리스트
        cnt_to_SRL=0                #Squad request list를 위한 변수
        for i in self.asked_squad_idx_list:
            cnt = 0
            self.Squad_Request_list[cnt_to_SRL].append(self.Squadlist[i])
            print(self.Squadlist[i]+" - ", end='')
            for j in self.sheet_request.columns:
                cnt= cnt+1
                if(cnt>self.showIdx and cnt<=self.showIdx+3 and j[i].value!=None):
                    self.Squad_Request_list[cnt_to_SRL].append(j[i].value)
                    print(j[i].value,end=' ')
            cnt_to_SRL = cnt_to_SRL + 1
            print()
        next_work = int(input('다음 작업을 선택해주세요 : 1-메인화면, 2-면제요청 사이클 반영 \n'))
        while (next_work < 1 or next_work > 2):
            next_work = int(input('입력 에러! 범위에 맞는 숫자 값을 입력해주세요 \n'))
        if (next_work == 1):
            self.Main_Page()
        else:
            self.Apply_Request()


    def Apply_Request(self):
        if(self.is_squad_selected==False):
            for i in range (len(self.asked_squad_idx_list)):
                if(i<len(self.asked_squad_idx_list)-1):
                    print(self.Squad_Request_list[i][0], end ='/')
                else:
                    print(self.Squad_Request_list[i][0], end=' ')
            print("중 요청을 적용시킬 분대를 소대-분대 형식으로 입력해주세요")
            Squad_to_Apply = input("(ex 1소대 1분대-> 1-1, 처음 화면으로 돌아가고 싶으면 0을 입력해주세요)\n").split('-')
            self.Squad_to_Apply_string = None
            if (Squad_to_Apply[0] == '0'):
                self.Main_Page()
            elif (int(Squad_to_Apply[0]) < 1 or int(Squad_to_Apply[0]) > 4 or int(Squad_to_Apply[1]) < 1 or int(Squad_to_Apply[1]) > 4 or len(Squad_to_Apply) == 1):
                while (int(Squad_to_Apply[0]) < 1 or int(Squad_to_Apply[0]) > 4 or int(Squad_to_Apply[1]) < 1 or int(Squad_to_Apply[1]) > 4 or len(Squad_to_Apply) == 1):
                    Squad_to_Apply = input("입력 오류. 다시 입력해주세요\n").split('-')
                self.Squad_to_Apply_string = Squad_to_Apply[0] + '소대 ' + Squad_to_Apply[1] + '분대'
            else:
                self.Squad_to_Apply_string = Squad_to_Apply[0] + '소대 ' + Squad_to_Apply[1] + '분대'

            while (self.Find_First_Index(self.asked_squad_list, self.Squad_to_Apply_string)== -1):
                Squad_to_Apply = input("입력 오류 : 면제를 요청한 분대들 중에서 선택해주세요. 처음 화면으로 돌아가고 싶으면 0을 입력해주세요)\n").split('-')
                if (Squad_to_Apply[0] == '0'):
                    self.Main_Page()
                elif (int(Squad_to_Apply[0]) < 1 or int(Squad_to_Apply[0]) > 4 or int(Squad_to_Apply[1]) < 1 or int(
                        Squad_to_Apply[1]) > 4 or len(Squad_to_Apply) == 1):
                    while (int(Squad_to_Apply[0]) < 1 or int(Squad_to_Apply[0]) > 4 or int(Squad_to_Apply[1]) < 1 or int(
                            Squad_to_Apply[1]) > 4 or len(Squad_to_Apply) == 1):
                        Squad_to_Apply = input("입력 오류. 다시 입력해주세요\n").split('-')
                    Squad_to_Apply_string = Squad_to_Apply[0] + '소대 ' + Squad_to_Apply[1] + '분대'
                else:
                    Squad_to_Apply_string = Squad_to_Apply[0] + '소대 ' + Squad_to_Apply[1] + '분대'

        worklist=list()
        for i in self.sheet_work.columns:
            worklist.append(i[self.Find_First_Index(self.Squadlist,self.Squad_to_Apply_string)].value)
        print(self.Squad_Request_list)
        print(worklist)
        change_squad_idx=self.Find_First_Index(self.Squadlist,self.Squad_to_Apply_string)
        change_work_idx=self.Find_First_Index(worklist,None)
        print('사역 사이클에 어떻게 표시할 지 내용을 입력해주세요.')
        content = input(' ex) o, 7/4일, OO사역 등등.. 메인으로 돌아가고 싶으면 0을 입력해주세요\n')
        if (content=="0"):
            self.Main_Page()
        if (change_work_idx != -1):
            self.sheet_work.cell(row=change_squad_idx + 1, column=change_work_idx + 1).value = content
            self.sheet_request.cell(row=change_squad_idx + 1, column=change_work_idx + 1).value =None
        else:
            self.sheet_work.cell(row=change_squad_idx + 1, column=len(worklist) + 1).value = content
            self.sheet_request.cell(row=change_squad_idx + 1, column=len(worklist) + 1).value = None

        self.wb.save('WorkCycle.xlsx')
        print("수정이 완료되었습니다.")


        next_work = int(input('다음 작업을 선택해주세요 : 0-종료, 1-메인화면, 2-같은 분대 면제권 적용, 3- 다른 분대 면제권 적용 \n'))
        while (next_work < 0 or next_work > 3):
            next_work = int(input('입력 에러! 범위에 맞는 숫자 값을 입력해주세요 \n'))
        if (next_work == 1):
            self.Main_Page()
        elif(next_work == 2):
            self.is_squad_selected=True
            self.Apply_Request()
        elif (next_work == 3):
            self.is_squad_selected=False
            self.Apply_Request()
        else:
            return 0


    def Add_manager(self):
        print("새로 추가될 관리자가 사용할 비밀번호를 입력해주세요 :", end='')
        password=input()
        password=self.Hash_String(password)
        while (self.Find_First_Index(self.pass_list,password)!=-1):
            print('이미 사용하고 있는 비밀번호 입니다. 다른 비밀번호를 입력해주세요.',end=' ')
            password = self.Hash_String(input('만약 메인 화면으로 돌아가고 싶다면 main을 입력해주세요\n'))
            if(password=='FAD58DE7366495DB4650CFEFAC2FCD61'):
                self.Main_Page()

        row_to_input = len(self.pass_list)
        if ((self.Find_First_Index(self.pass_list,None)!=-1) and (row_to_input >self.Find_First_Index(self.pass_list,None))):
            row_to_input=self.Find_First_Index(self.pass_list,None)
        self.passSheet.cell(row=row_to_input+1,column=1).value=password
        self.wbpass.save('pass.xlsx')
        print("새로운 관리자로 접속하려면 프로그램 종료 후 다시 실행해주세요!")
        next_work = int(input('다음 작업을 선택해주세요 : 0-종료, 1-메인화면 \n'))
        while (next_work < 0 or next_work > 1 ):
            next_work = int(input('입력 에러! 범위에 맞는 숫자 값을 입력해주세요 \n'))
        if (next_work == 1):
            self.Main_Page()
        else:
            return 0


    def Find_First_Index(self, list1: list, val):  # 리스트 중 특정 값 갖는 배열의 인덱스 반환
        for i in range(len(list1)):
            if list1[i] == val:
                return i
        return -1


    def Find_Inds_list(self, list1 : list, val):        #리스트 중 특정 값 갖는 배열의 인덱스를 리스트로 반환
        idx_list=list()
        for i in range(len(list1)):
            if list1[i]==val:
                idx_list.append(i)
        return idx_list

    def is_Val_in_list(self, list1 : list, val): #리스트 안에 값 있는지 확인하여 진위판단
        for i in range (len(list1)):
            if(list1[i]==val):
                return True
        return False

def main():
    Manager()
if __name__ =="__main__":
    main()