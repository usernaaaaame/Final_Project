import openpyxl



class Client ():
    def __init__(self):
        self.wb = openpyxl.load_workbook(r'WorkCycle.xlsx')
        self.ws_names = self.wb.sheetnames
        self.ws_names_to_show = list()
        for i in range(len(self.ws_names)):
            if (i % 2) == 0 :
                self.ws_names_to_show.append(self.ws_names[i])
        self.Main_Page()



    def Main_Page(self):
        self.is_squad_selected = False
        self.is_sheet_selected = False
        id_to_task = int(input('하고 싶은 작업을 선택하여 숫자로 입력해 주세요 : 0-종료, 1-예정 사역 확인, 2-사역 면제 요청\n'))
        while (id_to_task < 0 or id_to_task>2):
            id_to_task = int(input('입력 에러! 0에서 2사이의 값을 입력해 주세요 : 0-종료, 1-예정 사역 확인, 2-사역 면제 요청\n'))
        if id_to_task ==1 :
            self.Check()
        elif id_to_task==2 :
            self.Request()
        elif id_to_task==0 :
            return 0


    def Check(self):
        print("사역 목록 입니다 : ", end='')
        len_wsts=len(self.ws_names_to_show)
        for i in range(len_wsts):
            print(i+1,end='-')
            if i == len_wsts:
                print(self.ws_names_to_show[i])
            else:
                print(self.ws_names_to_show[i],end=' ')
        print()
        self.id_to_search = int(input('확인 하고 싶은 사역의 번호를 적어주세요\n'))
        while (self.id_to_search<1 or self.id_to_search>len_wsts):
            self.id_to_search = int(input('입력 에러! 범위에 맞는 숫자 값을 입력해주세요 \n'))
        sheet = self.wb[self.ws_names_to_show[self.id_to_search-1]]
        self.is_sheet_selected=True
        which_num = list()          #몇번째 회차인지 계산하기 위한 리스트
        for i in sheet.columns:
            which_num.append(i[17].value)
        row_to_search = self.Find_First_Index(which_num, None) #다 끝나지 않은 회차의 열
        if(row_to_search==-1):
            print("사역이 모두 돌았습니다. 관리자에게 문의 후 대기하여 주세요")
            return 0

        is_squad_did = list()                           #어떤 분대가 안했는지 값을 저장하기 위한 리스트
        is_squad_did_next = list()              #다음 회차에서 어떤 분대가 안했는지 값 저장 위한 리스트
        for i in sheet.rows:
            is_squad_did.append(i[row_to_search].value)
            is_squad_did_next.append(i[row_to_search+1].value)
        del is_squad_did[-1]
        del is_squad_did_next[-1]

        Not_worked_squad_idx = self.Find_Inds_list(is_squad_did, None)     #사역을 아직 하지 않은 분대의 인덱스 저장 리스트
        Not_worked_squad_idx_next = self.Find_Inds_list(is_squad_did_next, None)  #다음 회차에서 사역을 하지 않은 분대의 저장 리스트
        Squad_list=list()           #해당 시트에서의 분대 리스트 저장
        for i in sheet.rows:
            Squad_list.append(i[0].value)

        Not_worked_squadlist = list()           #사역 안한 분대 리스트
        Not_worked_squadlist_next = list()
        Not_worked_squad_val = list()           #find ind with list 사용 위한 리스트
        Not_worked_squad_val_next = list()

        for i in range(len(Not_worked_squad_idx)):
            Not_worked_squadlist.append(Squad_list[Not_worked_squad_idx[i]])
            Not_worked_squad_val.append(None)
        for i in range(len(Not_worked_squad_idx_next)):
            Not_worked_squadlist_next.append(Squad_list[Not_worked_squad_idx_next[i]])
            Not_worked_squad_val_next.append(None)

        print(self.ws_names_to_show[self.id_to_search - 1] + "에서 앞으로 사역할 5개 분대는 ")
        for i in range(5):      #사역 아직 안한 인원들 앞에서 5분대 뽑기
            if len(Not_worked_squad_idx)>=5:
                if(i==4):
                    print(Squad_list[Not_worked_squad_idx[i]]+ " 순서입니다")
                else:
                    print(Squad_list[Not_worked_squad_idx[i]], end=' - ')
            else:
                if (i<len(Not_worked_squad_idx)):
                    print(Squad_list[Not_worked_squad_idx[i]], end=' - ')
                else:
                    if (i==4):
                        print(Squad_list[Not_worked_squad_idx_next[i-len(Not_worked_squad_idx)]]+ " 순서입니다")
                    else:
                        print(Squad_list[Not_worked_squad_idx_next[i-len(Not_worked_squad_idx)]],end=' - ')
        print()
        print("자신의 분대의 다음 사역이 언제인지 알고 싶으면 소대-분대 형식으로 입력해주세요")
        self.My_Squad = input("(ex 1소대 1분대-> 1-1, 처음 화면으로 돌아가고 싶으면 0을, 사역 면제를 요청하고 싶으면 1을 입력해주세요)\n").split('-')
        if (self.My_Squad[0]=='0'):
            self.Main_Page()
        elif (len(self.My_Squad)==1 and self.My_Squad[0] == '1'):
            self.Request()
        elif (int(self.My_Squad[0])<1 or int(self.My_Squad[0])>4 or int(self.My_Squad[1])<1 or int(self.My_Squad[1])>4 or len(self.My_Squad)==1):
            while (int(self.My_Squad[0])<1 or int(self.My_Squad[0])>4 or int(self.My_Squad[1])<1 or int(self.My_Squad[1])>4 or len(self.My_Squad)==1) :
                self.My_Squad = input("입력 오류. 다시 입력해주세요\n").split('-')
            self.My_Squad_string=self.My_Squad[0]+'소대 '+self.My_Squad[1]+'분대'
            self.is_squad_selected= True
        else:
            self.My_Squad_string=self.My_Squad[0]+'소대 '+self.My_Squad[1]+'분대'
            self.is_squad_selected= True

        if self.Find_Ind_With_Lists(Not_worked_squadlist,self.My_Squad_string,Not_worked_squad_val,None)!=-1:
            print(self.My_Squad_string+"는 "+ self.ws_names_to_show[self.id_to_search-1]+"에서 다음 "+ str(self.Find_First_Index(Not_worked_squadlist,self.My_Squad_string)+1)+"번째, ",end='')
            print(str(len(Not_worked_squad_idx)+self.Find_First_Index(Not_worked_squadlist_next,self.My_Squad_string)+1)+"번째의 사역분대입니다.")
        elif (self.Find_Ind_With_Lists(Not_worked_squadlist,self.My_Squad_string,Not_worked_squad_val,None)==-1) and(self.Find_Ind_With_Lists(Not_worked_squadlist_next,self.My_Squad_string,Not_worked_squad_val_next,None)!=-1):
            print(self.My_Squad_string + "는 " + self.ws_names_to_show[self.id_to_search - 1] + "에서 다음 " + str(len(Not_worked_squad_idx) + self.Find_First_Index(Not_worked_squadlist_next,self.My_Squad_string)+1)+ "번째의 사역분대입니다.")
        else :
            print(self.My_Squad_string+"는 "+self.ws_names_to_show[self.id_to_search-1]+"에서 다음 "+str(len(Not_worked_squad_idx)+len(Not_worked_squad_idx_next))+"번 안의 사역분대에는 해당하지 않습니다.")

        next_work = int(input('다음 작업을 선택해주세요 : 0-종료, 1-메인화면, 2-사역 면제 요청 \n'))
        while (next_work<0 or next_work>2 ):
            next_work = int(input('입력 에러! 범위에 맞는 숫자 값을 입력해주세요 \n'))
        if(next_work==1):
            self.Main_Page()
        elif(next_work==2):
            self.Request()
        elif(next_work==0):
            return 0

    def Request(self):
        if(self.is_squad_selected==False):
            print("자신의 분대를 소대-분대 형식으로 입력해주세요")
            self.My_Squad = input("(ex 1소대 1분대-> 1-1, 처음 화면으로 돌아가고 싶으면 0을 입력해주세요)\n").split('-')
            if (self.My_Squad[0] == '0'):
                self.Main_Page()
            elif (int(self.My_Squad[0]) < 1 or int(self.My_Squad[0]) > 4 or int(self.My_Squad[1]) < 1 or int(self.My_Squad[1]) > 4 or len(self.My_Squad) == 1):
                while (int(self.My_Squad[0]) < 1 or int(self.My_Squad[0]) > 4 or int(self.My_Squad[1]) < 1 or int(self.My_Squad[1]) > 4 or len(self.My_Squad) == 1):
                    self.My_Squad = input("입력 오류. 다시 입력해주세요\n").split('-')
                self.My_Squad_string = self.My_Squad[0] + '소대 ' + self.My_Squad[1] + '분대'
                self.is_squad_selected = True
            else:
                self.My_Squad_string = self.My_Squad[0] + '소대 ' + self.My_Squad[1] + '분대'
                self.is_squad_selected = True

        if(self.is_sheet_selected == False):
            print("사역 목록 입니다 : ", end='')
            len_wsts = len(self.ws_names_to_show)
            for i in range(len_wsts):
                print(i + 1, end='-')
                if i == len_wsts:
                    print(self.ws_names_to_show[i])
                else:
                    print(self.ws_names_to_show[i], end=' ')
            print()
            self.id_to_search = int(input('면제를 신청하고 싶은 사역의 번호를 적어주세요\n'))
            while (self.id_to_search < 1 or self.id_to_search > len_wsts):
                self.id_to_search = int(input('입력 에러! 범위에 맞는 숫자 값을 입력해주세요 \n'))
        sheet = self.wb[self.ws_names_to_show[self.id_to_search - 1]]
        self.is_sheet_selected = True
        sheet_to_request= self.wb[self.ws_names_to_show[self.id_to_search - 1]+"(수정요청)"] # 수정요청하는 페이지에 저장하기 위한 변수
        Squad_list = list()  # 해당 시트에서의 분대 리스트 저장
        for i in sheet.rows:
            Squad_list.append(i[0].value)
        My_squad_idx = self.Find_First_Index(Squad_list,self.My_Squad_string)
        Work_list=list()
        for i in sheet.columns:
            Work_list.append(i[My_squad_idx].value)
        print('면제 사유를 입력해주세요.')
        reason=input('ex)환경 사역 면제권, 중대 단결 행사에서 받은 면제권. 메인으로 돌아가고 싶으면 0을 입력해주세요\n')
        if (reason=="0"):
            self.Main_Page()
        if(self.Find_First_Index(Work_list,None)!=-1):
            sheet_to_request.cell(row=My_squad_idx+1 ,column=self.Find_First_Index(Work_list,None)+1).value=reason
        else:
            sheet_to_request.cell(row=My_squad_idx + 1,column=len(Work_list) + 1).value = reason
        print("요청이 전송 되었습니다. 확인까지 시간이 다소 소요될 수 있습니다.")
        self.wb.save('WorkCycle.xlsx')
        next_work = int(input('다음 작업을 선택해주세요 : 0-종료, 1-메인화면 \n'))
        while (next_work < 0 or next_work > 1 ):
            next_work = int(input('입력 에러! 범위에 맞는 숫자 값을 입력해주세요 \n'))
        if (next_work == 1):
            self.Main_Page()
        else:
            return 0

    def Find_First_Index(self,list1 : list, val):   #리스트 중 특정 값 갖는 배열의 인덱스 반환
        for i in range(len(list1)):
            if list1[i]==val:
                return i
        return -1

    def Find_Ind_With_Lists(self, list1: list , val1, list2: list, val2 ): #리스트 들과 값 매치되는 인덱스 반환
        for i in range(len(list1)):
            if list1[i]==val1 and list2[i]==val2 :
                return i
        return -1


    def Find_Inds_list(self, list1 : list, val):        #리스트 중 특정 값 갖는 배열의 인덱스를 리스트로 반환
        idx_list=list()
        for i in range(len(list1)):
            if list1[i]==val:
                idx_list.append(i)
        return idx_list


def main():
    Client()
if __name__ =="__main__":
    main()