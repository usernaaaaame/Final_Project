import pandas as pd
import openpyxl


wb = openpyxl.load_workbook('sample.xlsx')
ws_names = wb.sheetnames
ws_names_to_show = list()
for i in range(len(ws_names)):
    if (i % 2) == 0:
        ws_names_to_show.append(ws_names[i])
sheet = wb[ws_names_to_show[0]]
sheet_to_save = wb[ws_names_to_show[0]+"(수정요청)"]
print(sheet)
Squad_list = list()  # 해당 시트에서의 분대 리스트 저장
for i in sheet.rows:
    Squad_list.append(i[0].value)
print(Squad_list)
result=list()
b=None
for i in sheet.columns:
    result.append(i[6].value)
gap=list()
for i in range(len(result)):
    if(result[i]==b):
        gap.append(i)
print(result)
print(gap)
print(gap[0])
sheet_to_save.cell(row=7, column=gap[0]+1).value="test2"
wb.save('sample.xlsx')


# # 열 출력 예시
# cnt=0
# for i in sheet.rows:
#     print(i[0].value,end=' ')
#     print(i[1].value)
#     if (cnt>0):
#         gap.append((i[1].value))
#     cnt=+1
# del gap[-1]

#find fir ind 테스트
# result=list()
# list1=['o','a',None,'a','a']
# b='a'
# for i in range(len(list1)):
#     if list1[i]==b:
#         result.append(i)
#         print(i)
# print(result)

# find idx Lists 테스트
# list1 = ["가","나","다","나"]
# list2 = ["o",None,"o","o"]
# val1="라"
# val2="X"
# for i in range(len(list1)):
#     if list1[i]==val1 and list2[i]==val2 :
#         print(i)
#
# idx_list=list()
# for i in range(len(list1)):
#     if list1[i]==val1:
#         idx_list.append(i)
# print(idx_list)
# print(len(idx_list))